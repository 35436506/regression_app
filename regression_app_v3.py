import streamlit as st
import pandas as pd
import numpy as np
import io
import warnings
warnings.filterwarnings('ignore')

# ── Lazy-load heavy libraries once per app lifecycle ─────────────────────────
@st.cache_resource(show_spinner=False)
def _load_heavy_libs():
    import statsmodels.api as _sm
    from scipy import stats as _stats
    import matplotlib as _mpl
    _mpl.use('Agg')
    import matplotlib.pyplot as _plt
    import matplotlib.gridspec as _gs
    import seaborn as _sns
    return _sm, _stats, _mpl, _plt, _gs, _sns

_sm_mod, stats, _mpl_mod, plt, gridspec, sns = _load_heavy_libs()
sm = _sm_mod   # alias kept for rest of code

# ── sklearn (always available in requirements) ────────────────────────────────
@st.cache_resource(show_spinner=False)
def _load_sklearn():
    try:
        from sklearn.linear_model import Ridge, RidgeCV
        from sklearn.preprocessing import StandardScaler
        from sklearn.metrics import r2_score as _sk_r2
        return Ridge, RidgeCV, StandardScaler, _sk_r2, True
    except ImportError:
        return None, None, None, None, False

Ridge, RidgeCV, StandardScaler, _sk_r2, _SKLEARN_OK = _load_sklearn()

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Regression Analyst",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=DM+Sans:wght@300;400;500;700&display=swap');

html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
.stApp { background: linear-gradient(135deg, #0d1117 0%, #161b22 100%); color: #e6edf3; }
h1,h2,h3 { font-family: 'Space Mono', monospace; color: #e6edf3; }

[data-testid="stMetricValue"] { color: #e6edf3 !important; }
[data-testid="stMetricLabel"] { color: #8b949e !important; }
.stDataFrame td, .stDataFrame th { color: #e6edf3 !important; background: #161b22 !important; }
.stSelectbox div[data-baseweb="select"] { background: #161b22 !important; color: #e6edf3 !important; }
.stMultiSelect div[data-baseweb="select"] { background: #161b22 !important; color: #e6edf3 !important; }
div[data-baseweb="option"] { background: #161b22 !important; color: #e6edf3 !important; }
div[data-baseweb="popover"] { background: #161b22 !important; }
.stTextInput input, .stTextArea textarea { color: #e6edf3 !important; background: #161b22 !important; }
div[data-testid="stSidebar"] { background: #161b22 !important; border-right: 1px solid #30363d; }

.hero-title {
    font-family:'Space Mono',monospace; font-size:2.2rem; font-weight:700;
    background:linear-gradient(90deg,#58a6ff,#bc8cff,#f778ba);
    -webkit-background-clip:text; -webkit-text-fill-color:transparent; line-height:1.2;
}
.hero-sub { color:#8b949e; font-size:1rem; margin-bottom:1.5rem; }

.section-hdr {
    font-family:'Space Mono',monospace; font-size:0.72rem; text-transform:uppercase;
    letter-spacing:2px; color:#58a6ff; margin-bottom:0.8rem;
    border-bottom:1px solid #21262d; padding-bottom:0.5rem;
}
.card { background:#161b22; border:1px solid #30363d; border-radius:12px; padding:1.2rem 1.4rem; margin-bottom:1rem; }
.card-accent { border-left:4px solid #58a6ff; }

.badge { display:inline-block; padding:2px 10px; border-radius:20px; font-size:0.72rem;
         font-weight:600; font-family:'Space Mono',monospace; margin-right:4px; }
.badge-blue   { background:#1f3a5f; color:#58a6ff; }
.badge-green  { background:#1a3a2a; color:#3fb950; }
.badge-yellow { background:#3a2d10; color:#d29922; }
.badge-red    { background:#3d1f1f; color:#f85149; }
.badge-purple { background:#2d1f5f; color:#bc8cff; }

.warn-box { background:#2a1f0a; border:1px solid #d29922; border-radius:8px; padding:0.9rem 1.1rem; margin:0.5rem 0; color:#d29922; font-size:0.88rem; }
.ok-box   { background:#0a2a14; border:1px solid #3fb950; border-radius:8px; padding:0.9rem 1.1rem; margin:0.5rem 0; color:#3fb950; font-size:0.88rem; }
.err-box  { background:#2a0a0a; border:1px solid #f85149; border-radius:8px; padding:0.9rem 1.1rem; margin:0.5rem 0; color:#f85149; font-size:0.88rem; }
.info-box { background:#0a1a2a; border:1px solid #58a6ff; border-radius:8px; padding:0.9rem 1.1rem; margin:0.5rem 0; color:#a8d8ff; font-size:0.88rem; }

.interpret-box {
    background:#1c2333; border:1px solid #58a6ff; border-radius:10px;
    padding:1.1rem 1.3rem; margin:0.8rem 0; color:#e6edf3;
    font-size:0.88rem; line-height:1.8;
}

.stButton>button {
    background:linear-gradient(90deg,#1f3a5f,#2d4a7a); color:#58a6ff; border:1px solid #58a6ff;
    border-radius:8px; font-family:'Space Mono',monospace; font-weight:700;
    padding:0.5rem 1.2rem; transition:all 0.2s;
}
.stButton>button:hover { background:linear-gradient(90deg,#2d4a7a,#3a5a9a); opacity:0.9; }

.run-btn>button {
    background:linear-gradient(90deg,#238636,#2ea043) !important; color:white !important;
    border:none !important;
}

div[data-testid="stExpander"] { background:#161b22; border:1px solid #30363d; border-radius:8px; }
</style>
""", unsafe_allow_html=True)

# ── Session state defaults ─────────────────────────────────────────────────
DEFAULTS = {
    "df": None,
    "filename": "",
    "header_row": 0,
    "run_history": [],
    "run_counter": 0,
    "outlier_rows": [],
    "excluded_rows": set(),
    "outlier_checked": False,
}
for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── Helper utilities ───────────────────────────────────────────────────────

def compute_vif(df_model, ind_vars):
    """
    Compute Variance Inflation Factor for each predictor.
    Returns DataFrame with columns: Biến, VIF, Mức độ
    Also returns a bool: has_high_vif (VIF > 10 for any variable)
    """
    from statsmodels.stats.outliers_influence import variance_inflation_factor
    X_df = df_model[ind_vars].dropna().copy()
    X_const = sm.add_constant(X_df)
    vif_data = []
    cols = list(X_const.columns)
    for i, col in enumerate(cols):
        if col == 'const':
            continue
        try:
            vif_val = variance_inflation_factor(X_const.values, i)
        except Exception:
            vif_val = np.nan
        if np.isnan(vif_val) or np.isinf(vif_val):
            level = "Không xác định"
            color = "#8b949e"
        elif vif_val < 5:
            level = "✅ Tốt (< 5)"
            color = "#3fb950"
        elif vif_val < 10:
            level = "⚠️ Trung bình (5–10)"
            color = "#d29922"
        else:
            level = "🚨 Cao (> 10) — Đa cộng tuyến!"
            color = "#f85149"
        vif_data.append({"Biến": col, "VIF": round(float(vif_val), 2), "Mức độ": level, "_color": color})

    has_high_vif = any(v["VIF"] > 10 for v in vif_data if not (np.isnan(v["VIF"]) or np.isinf(v["VIF"])))

    # Detect if high VIF is likely due to polynomial (X and X²)
    poly_collision = False
    base_vars = set()
    sq_vars = set()
    for v in vif_data:
        name = v["Biến"]
        if name.endswith("²") or name.endswith("_Sq") or name.endswith("2") or "Xc2" in name or "Cust_Sq" in name:
            sq_vars.add(name)
        else:
            base_vars.add(name)
    # Check if any base var approximately matches a square var
    for bv in base_vars:
        for sv in sq_vars:
            if bv.lower() in sv.lower() or sv.lower().replace("²","").replace("_sq","").replace("2","").strip() in bv.lower():
                poly_collision = True
                break

    return pd.DataFrame(vif_data), has_high_vif, poly_collision


@st.cache_data(show_spinner=False)
def detect_outlier_rows(df):
    """
    Detect statistical outlier rows (aggregate/total rows are already removed at import).
    Returns list of dicts: {index, label, reason, severity}
    - Statistical outliers via IQR (> 3*IQR) -> severity='extreme'
    """
    suspects = []
    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    text_cols = df.select_dtypes(exclude=[np.number]).columns.tolist()

    # --- Statistical extreme outliers only (aggregate rows removed at import) ----
    already_flagged = set()
    for col in numeric_cols:
        col_data = df[col].dropna()
        if len(col_data) < 4:
            continue
        Q1 = col_data.quantile(0.25)
        Q3 = col_data.quantile(0.75)
        IQR = Q3 - Q1
        if IQR == 0:
            continue
        fence = 3.0 * IQR
        for idx in df.index:
            if idx in already_flagged:
                continue
            val = df.at[idx, col]
            if pd.notna(val) and (val < Q1 - fence or val > Q3 + fence):
                # build label from first text column or index
                label_parts = []
                for tc in text_cols[:2]:
                    label_parts.append(f"{tc}={df.at[idx, tc]}")
                label = ', '.join(label_parts) if label_parts else f"row {idx}"
                direction = 'cao bất thường' if val > Q3 + fence else 'thấp bất thường'
                suspects.append({
                    'index': idx,
                    'label': label,
                    'reason': f"{col} = {val:,.0f} — {direction} (ngoài 3×IQR)",
                    'severity': 'extreme'
                })
                already_flagged.add(idx)

    return suspects


def detect_header_row(raw_df, max_scan=10):
    """
    Scan the first `max_scan` rows to find the true header row.
    The header row is the first row where:
      - Most cells are non-empty strings (not numbers, not NaN)
      - The row below contains predominantly numeric values
    Returns the header row index (0-based).
    """
    n_cols = raw_df.shape[1]
    scan = min(max_scan, raw_df.shape[0] - 1)

    for i in range(scan):
        row = raw_df.iloc[i]
        # Count cells that look like text labels (not numeric, not null)
        text_count = sum(
            1 for v in row
            if v is not None
            and not (isinstance(v, float) and np.isnan(v))
            and not isinstance(v, (int, float, np.integer, np.floating))
        )
        text_ratio = text_count / n_cols if n_cols > 0 else 0

        # Check that the NEXT row is mostly numeric
        if i + 1 < raw_df.shape[0]:
            next_row = raw_df.iloc[i + 1]
            num_count = sum(1 for v in next_row if isinstance(v, (int, float, np.integer, np.floating))
                            and not (isinstance(v, float) and np.isnan(v)))
            num_ratio = num_count / n_cols if n_cols > 0 else 0
        else:
            num_ratio = 0

        if text_ratio >= 0.5 and num_ratio >= 0.4:
            return i

    return 0   # fallback: row 0 is the header




@st.cache_data(show_spinner="Đang đọc file...")
def _load_data_core(file_bytes: bytes, file_name: str, chosen_sheet=None):
    """Pure cached file reader — no Streamlit side-effects."""
    name = file_name.lower()
    header_row = 0

    if name.endswith(".csv"):
        buf = io.BytesIO(file_bytes)
        try:
            raw = pd.read_csv(buf, header=None, nrows=15)
        except Exception:
            buf = io.BytesIO(file_bytes)
            raw = pd.read_csv(buf, header=None, nrows=15, encoding='latin1')
        header_row = detect_header_row(raw)
        buf = io.BytesIO(file_bytes)
        try:
            df = pd.read_csv(buf, header=header_row)
        except Exception:
            buf = io.BytesIO(file_bytes)
            df = pd.read_csv(buf, header=header_row, encoding='latin1')

    elif name.endswith((".xlsx", ".xls")):
        buf = io.BytesIO(file_bytes)
        xf = pd.ExcelFile(buf)
        sheet = chosen_sheet or xf.sheet_names[0]
        buf = io.BytesIO(file_bytes)
        raw = pd.read_excel(buf, sheet_name=sheet, header=None, nrows=15)
        header_row = detect_header_row(raw)
        buf = io.BytesIO(file_bytes)
        df = pd.read_excel(buf, sheet_name=sheet, header=header_row)
    else:
        return None, 0, []

    df = df.dropna(how='all').reset_index(drop=True)
    df.columns = [str(c).strip() for c in df.columns]

    AGG_KW = {'total', 'tong', 'tổng', 'sum', 'grand total', 'subtotal', 'cộng', 'grand'}
    text_cols = df.select_dtypes(exclude=['number']).columns.tolist()
    removed_labels = []
    if len(df) > 0 and text_cols:
        last_idx = df.index[-1]
        if any(str(df.at[last_idx, tc]).strip().lower() in AGG_KW for tc in text_cols):
            removed_labels = [str(df.at[last_idx, text_cols[0]])]
            df = df.iloc[:-1].reset_index(drop=True)

    for col in df.columns:
        if df[col].dtype == object:
            converted = pd.to_numeric(df[col], errors='coerce')
            if converted.notna().sum() / max(len(df), 1) > 0.7:
                df[col] = converted

    return df, header_row, removed_labels


def load_data(uploaded_file):
    """Thin wrapper: handles sheet-selection UI then calls cached core."""
    name = uploaded_file.name.lower()
    file_bytes = uploaded_file.read()

    chosen_sheet = None
    if name.endswith((".xlsx", ".xls")):
        xf = pd.ExcelFile(io.BytesIO(file_bytes))
        sheet_names = xf.sheet_names
        if len(sheet_names) > 1:
            chosen_sheet = st.sidebar.selectbox("Sheet name", sheet_names)
        else:
            chosen_sheet = sheet_names[0]

    result = _load_data_core(file_bytes, uploaded_file.name, chosen_sheet)
    if result[0] is None:
        st.error("Unsupported file type.")
        return None, 0

    df, header_row, removed_labels = result
    st.session_state["_agg_rows_removed"] = removed_labels
    return df, header_row



@st.cache_data(show_spinner=False)
def suggest_variables(df):
    """
    Heuristically suggest which column is the dependent variable (Y)
    and which are independent variables (X).

    Rules:
    - Y candidates: numeric columns whose name contains output-like keywords
      (price, cost, revenue, sales, profit, expense, score, output, result,
       doanh, chi, gia, ket, thu, buchanan, votes_for)
      OR the numeric column with the highest variance relative to mean (CV).
      Tie-break: last numeric column (common in textbook datasets).
    - X candidates: all other numeric columns.
    - Text/ID columns (high cardinality strings, index-like) are flagged separately.
    """
    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    if not numeric_cols:
        return None, [], [], "Không tìm thấy cột số nào."

    # Keywords that suggest an outcome/dependent variable
    Y_KEYWORDS = [
        "price", "cost", "revenue", "sales", "profit", "expense", "score",
        "output", "result", "income", "wage", "salary", "return", "loss",
        "target", "label", "y", "dependent", "outcome", "response",
        # Vietnamese
        "gia", "chi", "phi", "thu", "doanh", "ket", "qua", "luong",
        # Case-specific
        "buchanan", "vote", "expense", "carats",
    ]

    best_y = None
    best_score = -1
    reasons = {}

    for col in numeric_cols:
        col_lower = col.lower()
        score = 0
        reason_parts = []

        # Keyword match
        kw_matches = [kw for kw in Y_KEYWORDS if kw in col_lower]
        if kw_matches:
            score += 3
            reason_parts.append(f"tên gợi ý biến kết quả ({', '.join(kw_matches)})")

        # High CV (coefficient of variation) → likely outcome with wide range
        col_data = df[col].dropna()
        if col_data.mean() != 0:
            cv = col_data.std() / abs(col_data.mean())
            if cv > 1.5:
                score += 1
                reason_parts.append(f"hệ số biến thiên cao (CV={cv:.2f})")

        # Last column heuristic (common in structured datasets)
        if col == numeric_cols[-1]:
            score += 0.5
            reason_parts.append("cột số cuối cùng")

        reasons[col] = reason_parts if reason_parts else ["không có dấu hiệu rõ ràng"]
        if score > best_score:
            best_score = score
            best_y = col

    # If no keyword matched at all, default to last numeric col
    if best_y is None:
        best_y = numeric_cols[-1]

    best_x = [c for c in numeric_cols if c != best_y]

    # Build explanation
    y_reason = "; ".join(reasons.get(best_y, []))
    explanation = (
        f"🎯 **Gợi ý Y = `{best_y}`** — {y_reason}.\n\n"
        f"📌 **Gợi ý X = {', '.join([f'`{c}`' for c in best_x])}** — các biến số còn lại.\n\n"
        "Bạn có thể thay đổi lựa chọn bên dưới."
    )

    return best_y, best_x, numeric_cols, explanation


@st.cache_data(show_spinner=False)
def data_quality_report(df):
    issues = []
    numeric_df = df.select_dtypes(include=[np.number])

    # Missing values
    missing = df.isnull().sum()
    total_missing = missing.sum()
    if total_missing > 0:
        miss_cols = missing[missing > 0]
        pct = (miss_cols / len(df) * 100).round(1)
        issues.append({
            "level": "warn",
            "msg": f"Thiếu dữ liệu: {total_missing} giá trị null trong {len(miss_cols)} cột — " +
                   ", ".join([f"{c} ({p}%)" for c, p in pct.items()])
        })

    # Duplicate rows
    dupes = df.duplicated().sum()
    if dupes > 0:
        issues.append({"level": "warn", "msg": f"Hàng trùng lặp: {dupes} hàng bị lặp."})

    # Low-variance columns
    for col in numeric_df.columns:
        if numeric_df[col].std() == 0:
            issues.append({"level": "err", "msg": f"Cột '{col}' có phương sai = 0 (hằng số), không dùng được trong hồi quy."})

    # Outliers via IQR
    outlier_cols = []
    for col in numeric_df.columns:
        Q1 = numeric_df[col].quantile(0.25)
        Q3 = numeric_df[col].quantile(0.75)
        IQR = Q3 - Q1
        n_out = ((numeric_df[col] < Q1 - 1.5 * IQR) | (numeric_df[col] > Q3 + 1.5 * IQR)).sum()
        if n_out > 0:
            outlier_cols.append(f"{col} ({n_out} điểm)")
    if outlier_cols:
        issues.append({"level": "warn", "msg": "Outlier (IQR): " + ", ".join(outlier_cols)})

    # Non-numeric columns
    non_num = df.select_dtypes(exclude=[np.number]).columns.tolist()
    if non_num:
        issues.append({"level": "info", "msg": f"Cột phi số (cần encode hoặc loại bỏ trước khi hồi quy): {', '.join(non_num)}"})

    if not issues:
        issues.append({"level": "ok", "msg": "Dữ liệu sạch — không phát hiện vấn đề nghiêm trọng."})

    return issues


def get_ols_stats(model):
    reg_stats = pd.DataFrame({
        'Chỉ số': ['Multiple R', 'R²', 'Adjusted R²', 'Std Error', 'Observations'],
        'Giá trị': [
            float(np.sqrt(model.rsquared)),
            float(model.rsquared),
            float(model.rsquared_adj),
            float(np.sqrt(model.mse_resid)),
            int(model.nobs)
        ]
    })
    anova = pd.DataFrame({
        'ANOVA': ['Regression', 'Residual', 'Total'],
        'df': [int(model.df_model), int(model.df_resid), int(model.df_model + model.df_resid)],
        'SS': [model.ess, model.ssr, model.centered_tss],
        'MS': [model.ess / model.df_model, model.ssr / model.df_resid, np.nan],
        'F': [model.fvalue, np.nan, np.nan],
        'Significance F': [model.f_pvalue, np.nan, np.nan]
    })
    coef_df = pd.DataFrame({
        'Biến': model.params.index,
        'Hệ số': model.params.values,
        'Std Error': model.bse.values,
        't Stat': model.tvalues.values,
        'P-value': model.pvalues.values,
        'Lower 95%': model.conf_int()[0].values,
        'Upper 95%': model.conf_int()[1].values
    })
    return reg_stats, anova, coef_df


def build_X(df, ind_vars, model_type, x_col_for_quad=None):
    """Build design matrix based on model type."""
    if model_type == "Tuyến tính (Linear)":
        X = sm.add_constant(df[ind_vars])
    elif model_type == "Bậc hai (Quadratic)":
        X_df = df[ind_vars].copy()
        for col in ind_vars:
            X_df[col + "²"] = df[col] ** 2
        X = sm.add_constant(X_df)
    elif model_type == "Bậc hai Centered":
        x_col = ind_vars[0] if len(ind_vars) == 1 else (x_col_for_quad or ind_vars[0])
        rest = [c for c in ind_vars if c != x_col]
        x_mean = df[x_col].mean()
        X_df = pd.DataFrame({'Xc': df[x_col] - x_mean, 'Xc2': (df[x_col] - x_mean) ** 2})
        for c in rest:
            X_df[c] = df[c]
        X = sm.add_constant(X_df)
    elif model_type == "Logarithmic (Log Y)":
        y_check = None
        X = sm.add_constant(df[ind_vars])
        # y will be log-transformed outside
    elif model_type == "Logarithmic (Log X)":
        X_df = pd.DataFrame(index=df.index)
        for col in ind_vars:
            X_df[f"ln({col})"] = np.log(df[col])
        X = sm.add_constant(X_df)
    elif model_type == "Tương tác (Interaction)":
        X_df = df[ind_vars].copy()
        cols = list(ind_vars)
        for i in range(len(cols)):
            for j in range(i + 1, len(cols)):
                X_df[f"{cols[i]}×{cols[j]}"] = df[cols[i]] * df[cols[j]]
        X = sm.add_constant(X_df)
    else:
        X = sm.add_constant(df[ind_vars])
    return X


def run_regression(df, dep_var, ind_vars, model_type):
    df_clean = df[[dep_var] + ind_vars].dropna()
    y_raw = df_clean[dep_var]

    if model_type == "Logarithmic (Log Y)":
        if (y_raw <= 0).any():
            raise ValueError("Log Y yêu cầu tất cả giá trị Y > 0.")
        y = np.log(y_raw)
    elif model_type == "Logarithmic (Log X)":
        for col in ind_vars:
            if (df_clean[col] <= 0).any():
                raise ValueError(f"Log X yêu cầu tất cả giá trị của '{col}' > 0.")
        y = y_raw
    else:
        y = y_raw

    X = build_X(df_clean, ind_vars, model_type)
    model = sm.OLS(y, X).fit()
    return model, df_clean, y


def make_plots(model, df_clean, dep_var, ind_vars, model_type):
    """
    Generate 4 diagnostic plots.
    Plot 1 (Trend): always 2-D — X axis = most important variable by |t-stat|,
                    all other vars held at their mean. Prediction done over a
                    sorted linspace so the line is never zigzag.
    Plot 2: Residuals vs Fitted
    Plot 3: Q-Q plot
    Plot 4: Actual vs Predicted
    """
    # ── colour palette ──────────────────────────────────────────────────────
    DARK  = '#0d1117'
    PANEL = '#1c2333'
    GRID  = '#30363d'
    RED   = '#f85149'
    BLUE  = '#58a6ff'
    GRAY  = '#8b949e'
    GREEN = '#3fb950'
    WHITE = '#e6edf3'
    CI_COLOR = '#f85149'

    is_log = (model_type == "Logarithmic (Log Y)")
    is_log_x = (model_type == "Logarithmic (Log X)")
    is_interaction = (model_type == "Tương tác (Interaction)")

    # ── pick most important X: highest |t-stat| among ind_vars ──────────────
    t_abs = {}
    for col in ind_vars:
        # match partial name (quadratic terms contain the base name)
        for param_name in model.tvalues.index:
            if param_name == col or param_name.startswith(col):
                t_abs[col] = max(t_abs.get(col, 0), abs(model.tvalues[param_name]))
    if t_abs:
        key_var = max(t_abs, key=t_abs.get)
    else:
        key_var = ind_vars[0]

    # ── figure layout ────────────────────────────────────────────────────────
    fig = plt.figure(figsize=(16, 11))
    fig.patch.set_facecolor(DARK)
    gs = gridspec.GridSpec(2, 2, figure=fig, hspace=0.48, wspace=0.38)

    def style_ax(ax, title, xlabel, ylabel):
        ax.set_facecolor(PANEL)
        ax.set_title(title, color=WHITE, fontsize=11, fontweight='bold', pad=10)
        ax.set_xlabel(xlabel, color=GRAY, fontsize=9, labelpad=6)
        ax.set_ylabel(ylabel, color=GRAY, fontsize=9, labelpad=6)
        ax.tick_params(colors=GRAY, labelsize=8)
        ax.grid(True, alpha=0.25, color=GRID, linestyle='--')
        for spine in ax.spines.values():
            spine.set_color(GRID)

    # ════════════════════════════════════════════════════════════════════════
    # PLOT 1 — Trend (Interaction: multiple lines; others: single line)
    # ════════════════════════════════════════════════════════════════════════
    ax1 = fig.add_subplot(gs[0, 0])

    x_vals   = df_clean[key_var].values
    y_actual = df_clean[dep_var].values

    ax1.scatter(x_vals, y_actual, color=GRAY, alpha=0.55, s=35,
                zorder=3, label='Dữ liệu thực tế', edgecolors='none')

    x_grid = np.linspace(x_vals.min(), x_vals.max(), 300)

    if is_interaction and len(ind_vars) >= 2:
        # ── Interaction plot: draw one regression line per level of the
        #    moderator variable (ind_vars[1]).  We use p10 / mean / p90
        #    of the moderator so the lines span realistic values.
        # ────────────────────────────────────────────────────────────────
        mod_var   = ind_vars[1]          # moderator (second X)
        other_vars = [v for v in ind_vars if v not in (key_var, mod_var)]

        mod_vals   = df_clean[mod_var]
        unique_mod = sorted(mod_vals.unique())

        # ── Detect binary (0/1 dummy) vs continuous moderator ──────────────
        is_binary_mod = (
            set(unique_mod).issubset({0, 1}) or
            (len(unique_mod) == 2 and unique_mod[0] == 0 and unique_mod[1] == 1)
        )

        if is_binary_mod:
            # Binary moderator: draw exactly 2 lines — one per group
            mod_levels = {
                f"{mod_var} = 0  (nhóm gốc)": 0.0,
                f"{mod_var} = 1  (nhóm can thiệp)": 1.0,
            }
            LINE_COLORS = ['#58a6ff', '#3fb950']   # blue, green
            subtitle_note = (
                f"{mod_var} là biến nhị phân (0/1) → 2 đường cho 2 nhóm"
                + (f"  |  biến khác giữ ở mean" if other_vars else "")
            )
        else:
            # Continuous moderator: draw 3 lines at ±1SD and mean
            # (±1 SD is statistically standard — more interpretable than p10/p90)
            mod_mean = float(mod_vals.mean())
            mod_sd   = float(mod_vals.std())
            mod_levels = {
                f"{mod_var} = −1SD ({mod_mean - mod_sd:.2f})": mod_mean - mod_sd,
                f"{mod_var} = mean ({mod_mean:.2f})":          mod_mean,
                f"{mod_var} = +1SD ({mod_mean + mod_sd:.2f})": mod_mean + mod_sd,
            }
            LINE_COLORS = ['#58a6ff', '#f85149', '#3fb950']   # blue, red, green
            subtitle_note = (
                f"3 đường = −1SD/mean/+1SD của {mod_var}"
                + (f"  |  biến khác giữ ở mean" if other_vars else "")
            )

        for (lbl, mod_val), lc in zip(mod_levels.items(), LINE_COLORS):
            pred_rows = []
            for xv in x_grid:
                row = {c: float(df_clean[c].mean()) for c in other_vars}
                row[key_var] = xv
                row[mod_var] = mod_val
                pred_rows.append(row)
            df_grid = pd.DataFrame(pred_rows)
            X_grid  = build_X(df_grid, ind_vars, model_type)
            try:
                target_cols = model.model.exog_names
                for col in target_cols:
                    if col not in X_grid.columns:
                        X_grid[col] = 0.0
                X_grid = X_grid[target_cols]
                pf = model.get_prediction(X_grid).summary_frame(alpha=0.05)
                y_line = pf['mean'].values
                y_lo   = pf['mean_ci_lower'].values
                y_hi   = pf['mean_ci_upper'].values
                ax1.fill_between(x_grid, y_lo, y_hi, color=lc, alpha=0.10, zorder=2)
                ax1.plot(x_grid, y_line, color=lc, lw=2.2, zorder=4, label=lbl)
            except Exception:
                pass

        style_ax(ax1,
                 f"Interaction: {dep_var} ~ {key_var}  (theo mức {mod_var})",
                 key_var, dep_var)
        ax1.annotate(
            subtitle_note,
            xy=(0.5, -0.14), xycoords='axes fraction',
            ha='center', fontsize=7.5, color=GRAY)
        ax1.legend(fontsize=7.5, labelcolor=WHITE, facecolor=PANEL,
                   edgecolor=GRID, framealpha=0.9)

    else:
        # ── Standard single-line trend (Linear / Quadratic / Log Y / Log X) ──
        pred_rows = []
        for xv in x_grid:
            row = {c: float(df_clean[c].mean()) for c in ind_vars}
            row[key_var] = xv
            pred_rows.append(row)
        df_grid = pd.DataFrame(pred_rows)
        X_grid  = build_X(df_grid, ind_vars, model_type)
        try:
            target_cols = model.model.exog_names
            for col in target_cols:
                if col not in X_grid.columns:
                    X_grid[col] = 0.0
            X_grid = X_grid[target_cols]
        except Exception:
            pass

        try:
            preds = model.get_prediction(X_grid)
            pf    = preds.summary_frame(alpha=0.05)
            if is_log:
                y_line = np.exp(pf['mean'].values)
                y_lo   = np.exp(pf['mean_ci_lower'].values)
                y_hi   = np.exp(pf['mean_ci_upper'].values)
            else:
                y_line = pf['mean'].values
                y_lo   = pf['mean_ci_lower'].values
                y_hi   = pf['mean_ci_upper'].values

            ax1.fill_between(x_grid, y_lo, y_hi,
                             color=CI_COLOR, alpha=0.18, zorder=2, label='95% CI')
            ax1.plot(x_grid, y_line, color=RED, lw=2.5, zorder=4, label='Đường hồi quy')
        except Exception:
            sort_idx = np.argsort(x_vals)
            y_fit = np.exp(model.fittedvalues.values) if is_log else model.fittedvalues.values
            ax1.plot(x_vals[sort_idx], y_fit[sort_idx], color=RED, lw=2.5, label='Fitted')

        other_vars = [v for v in ind_vars if v != key_var]
        note = ""
        if other_vars:
            means_str = ", ".join([f"{v}={df_clean[v].mean():.2f}" for v in other_vars])
            note = f"(biến khác giữ ở mean: {means_str})"
        if is_log_x:
            note = f"Trục X gốc — hồi quy trên ln({key_var})" + (f"  |  {note}" if note else "")

        title_main = f"Trend: {dep_var} ~ {key_var}"
        if len(ind_vars) > 1:
            title_main += "  [biến quan trọng nhất]"
        style_ax(ax1, title_main, key_var, dep_var)
        if note:
            ax1.annotate(note, xy=(0.5, -0.14), xycoords='axes fraction',
                         ha='center', fontsize=7.5, color=GRAY)
        ax1.legend(fontsize=8.5, labelcolor=WHITE, facecolor=PANEL,
                   edgecolor=GRID, framealpha=0.9)

    # ════════════════════════════════════════════════════════════════════════
    # PLOT 2 — Residuals vs Fitted
    # ════════════════════════════════════════════════════════════════════════
    ax2 = fig.add_subplot(gs[0, 1])
    fitted = model.fittedvalues.values
    resid  = model.resid.values
    ax2.scatter(fitted, resid, color=BLUE, alpha=0.55, s=40,
                edgecolors='none', zorder=3)
    ax2.axhline(0, color=RED, linestyle='--', lw=1.8, zorder=2)
    # Lowess smoother hint
    try:
        from statsmodels.nonparametric.smoothers_lowess import lowess
        sm_line = lowess(resid, fitted, frac=0.5)
        ax2.plot(sm_line[:, 0], sm_line[:, 1], color=GREEN,
                 lw=1.5, linestyle='-', alpha=0.7, label='Lowess')
        ax2.legend(fontsize=8, labelcolor=WHITE, facecolor=PANEL, edgecolor=GRID)
    except Exception:
        pass
    style_ax(ax2, 'Residual Plot', 'Giá trị dự báo (Fitted)', 'Phần dư (Residuals)')

    # ════════════════════════════════════════════════════════════════════════
    # PLOT 3 — Q-Q Plot
    # ════════════════════════════════════════════════════════════════════════
    ax3 = fig.add_subplot(gs[1, 0])
    (osm, osr), (slope, intercept, _) = stats.probplot(resid, dist="norm")
    ax3.scatter(osm, osr, color=BLUE, alpha=0.65, s=35, edgecolors='none',
                zorder=3, label='Phần dư')
    x_ref = np.array([min(osm), max(osm)])
    ax3.plot(x_ref, slope * x_ref + intercept, color=RED,
             lw=2, zorder=4, label='Đường chuẩn')
    style_ax(ax3, 'Q-Q Plot  (Chuẩn hóa phần dư)',
             'Quantile lý thuyết', 'Quantile thực tế')
    ax3.legend(fontsize=8.5, labelcolor=WHITE, facecolor=PANEL, edgecolor=GRID)

    # ════════════════════════════════════════════════════════════════════════
    # PLOT 4 — Actual vs Predicted
    # ════════════════════════════════════════════════════════════════════════
    ax4 = fig.add_subplot(gs[1, 1])
    # Log Y → back-transform predictions; Log X → y is on original scale
    y_pred_p = np.exp(model.fittedvalues.values) if is_log else model.fittedvalues.values
    y_act_p  = np.exp(df_clean[dep_var].values) if is_log else df_clean[dep_var].values
    ax4.scatter(y_act_p, y_pred_p, color=GREEN, alpha=0.55, s=40,
                edgecolors='none', zorder=3)
    mn = min(y_act_p.min(), y_pred_p.min())
    mx = max(y_act_p.max(), y_pred_p.max())
    ax4.plot([mn, mx], [mn, mx], color=RED, lw=1.8, linestyle='--',
             zorder=4, label='Perfect fit (y=x)')
    style_ax(ax4, 'Actual vs Predicted', 'Giá trị thực tế', 'Giá trị dự báo')
    ax4.legend(fontsize=8.5, labelcolor=WHITE, facecolor=PANEL, edgecolor=GRID)

    # ── overall title ────────────────────────────────────────────────────────
    n_vars_label = f"{len(ind_vars)} biến: {', '.join(ind_vars)}" if len(ind_vars) > 1 else ind_vars[0]
    fig.suptitle(
        f"{dep_var}  ←  {n_vars_label}   [{model_type}]   "
        f"R²={model.rsquared:.3f}  Adj R²={model.rsquared_adj:.3f}",
        color=WHITE, fontsize=11, fontweight='bold', y=1.01
    )

    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=140, bbox_inches='tight', facecolor=DARK)
    plt.close()
    buf.seek(0)
    return buf


def interpret_model(model, model_type, dep_var, ind_vars):
    """Generate Vietnamese interpretation text."""
    r2 = model.rsquared
    adj_r2 = model.rsquared_adj
    f_p = model.f_pvalue
    n = int(model.nobs)
    k = int(model.df_model)
    rmse = float(np.sqrt(model.mse_resid))

    lines = []

    # Overall model fit
    lines.append("📊 ĐÁNH GIÁ TỔNG THỂ MÔ HÌNH")
    if r2 >= 0.9:
        fit_quality = "rất tốt"
    elif r2 >= 0.7:
        fit_quality = "tốt"
    elif r2 >= 0.5:
        fit_quality = "trung bình"
    else:
        fit_quality = "yếu"
    lines.append(f"• R² = {r2:.4f} → Mô hình giải thích được {r2*100:.1f}% biến động của {dep_var}. Độ phù hợp {fit_quality}.")
    lines.append(f"• Adjusted R² = {adj_r2:.4f} (sau khi điều chỉnh số biến k={k}).")

    if f_p < 0.001:
        lines.append(f"• F-test p-value < 0.001 → Mô hình có ý nghĩa thống kê rất cao (***)")
    elif f_p < 0.05:
        lines.append(f"• F-test p-value = {f_p:.4f} → Mô hình có ý nghĩa thống kê (**).")
    else:
        lines.append(f"• F-test p-value = {f_p:.4f} → Mô hình KHÔNG có ý nghĩa thống kê (p ≥ 0.05).")

    lines.append(f"• RMSE = {rmse:.4f} (sai số dự báo trung bình).")
    lines.append("")

    # Coefficients
    lines.append("🔢 PHÂN TÍCH HỆ SỐ HỒI QUY")
    for var, coef, pval in zip(model.params.index, model.params.values, model.pvalues.values):
        if var == "const":
            lines.append(f"• Hằng số (const) = {coef:.4f}.")
            continue
        sig = "***" if pval < 0.001 else ("**" if pval < 0.01 else ("*" if pval < 0.05 else "⚠ không có ý nghĩa"))
        direction = "tăng" if coef > 0 else "giảm"
        lines.append(f"• {var}: hệ số = {coef:.4f}, p = {pval:.4f} ({sig}) → khi {var} tăng 1 đơn vị, {dep_var} {direction} {abs(coef):.4f} đơn vị (các biến khác không đổi).")

    lines.append("")

    # Residual diagnostics hint
    lines.append("🔍 GỢI Ý CHẨN ĐOÁN")
    if adj_r2 < r2 - 0.05 and k > 2:
        lines.append("• Adjusted R² thấp hơn R² nhiều → có thể có biến thừa, cân nhắc loại bỏ biến không có ý nghĩa.")
    if rmse > (model.model.endog.std() * 0.5):
        lines.append("• RMSE tương đối cao so với độ lệch chuẩn Y → sai số dự báo còn lớn.")
    if model_type in ["Bậc hai (Quadratic)", "Bậc hai Centered"]:
        lines.append("• Mô hình bậc 2: kiểm tra hệ số bậc 2 có ý nghĩa không. Nếu không → quay về mô hình tuyến tính.")
    if model_type == "Logarithmic (Log X)":
        lines.append("• Log X: hệ số β nghĩa là khi X tăng 1%, Y thay đổi β/100 đơn vị (quan hệ giảm dần — diminishing returns).")
        lines.append("• Kiểm tra: nếu phần dư phân bố đều → Log X phù hợp. Nếu vẫn còn hình phễu → thử Log Y.")
    lines.append("• Kiểm tra biểu đồ Residual: phần dư nên phân bố ngẫu nhiên quanh 0.")
    lines.append("• Q-Q Plot: các điểm gần đường thẳng → phần dư xấp xỉ chuẩn, giả thuyết OLS được thỏa.")

    return "\n".join(lines)


def export_to_excel(run_history):
    """Export all run results to Excel with detailed sheets (openpyxl engine)."""
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    # ── Helper: make styles ────────────────────────────────────────────────
    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color.lstrip("#"))

    def _font(bold=False, color="000000", name="Arial", size=10):
        return Font(bold=bold, color=color.lstrip("#"), name=name, size=size)

    def _border():
        s = Side(style="thin", color="AAAAAA")
        return Border(left=s, right=s, top=s, bottom=s)

    def _align(h="left", wrap=False):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    # pre-built style tuples  (fill, font, alignment, border)
    SUM_HDR  = (_fill("1f3a5f"), _font(bold=True, color="58a6ff"), _align("center"), _border())
    SUM_NUM  = (None,             _font(),                          _align("right"),  _border())
    SUM_INT  = (None,             _font(),                          _align("right"),  _border())
    SUM_TXT  = (None,             _font(),                          _align("left"),   _border())
    SEC_HDR  = (_fill("1f3a5f"), _font(bold=True, color="FFFFFF"), _align("center"), _border())
    COL_HDR  = (_fill("BDD7EE"), _font(bold=True, color="000000"), _align("center"), _border())
    RUN_NUM  = (None,             _font(color="000000"),            _align("right"),  _border())
    RUN_INT  = (None,             _font(color="000000"),            _align("right"),  _border())
    RUN_TXT  = (None,             _font(color="000000"),            _align("left"),   _border())
    INTERP_H = (_fill("1f3a5f"), _font(bold=True, color="FFFFFF"), _align("left"),   _border())
    INTERP_T = (None,             _font(color="000000"),            _align("left", wrap=True), None)

    def _apply(cell, style_tuple):
        fill, font, align, border = style_tuple
        if fill:   cell.fill      = fill
        if font:   cell.font      = font
        if align:  cell.alignment = align
        if border: cell.border    = border

    NUM_FMT  = '#,##0.0000'
    INT_FMT  = '0'

    # ── Summary sheet ──────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    summary_rows = []
    for r in run_history:
        rr       = r.get("ridge_result")
        is_ridge = r.get("is_ridge_run", False)
        fstat_v  = "" if is_ridge or (isinstance(r['fstat'], float) and np.isnan(r['fstat'])) else r['fstat']
        fp_v     = "" if is_ridge or (isinstance(r['f_pvalue'], float) and np.isnan(r['f_pvalue'])) else r['f_pvalue']
        adj_r2_v = "" if is_ridge else r['adj_r2']
        summary_rows.append({
            'Run #':           int(r['run_id']),
            'Loại':            'Ridge' if is_ridge else 'OLS',
            'Biến phụ thuộc':  r['dep_var'],
            'Biến độc lập':    ', '.join(r['ind_vars']),
            'Loại mô hình':    r['model_type'],
            'N':               int(r['n']),
            'R²':              r['r2'],
            'Adj R²':          adj_r2_v,
            'RMSE':            r['rmse'],
            'F-stat':          fstat_v,
            'F p-value':       fp_v,
            'Ridge λ':         rr['best_lambda'] if rr else '',
        })

    sum_df        = pd.DataFrame(summary_rows)
    col_widths    = [8, 8, 22, 32, 28, 8, 13, 13, 13, 13, 15, 11]
    int_sum_cols  = {'Run #', 'N'}

    for ci, (col, w) in enumerate(zip(sum_df.columns, col_widths), start=1):
        cell = ws_sum.cell(row=1, column=ci, value=col)
        _apply(cell, SUM_HDR)
        ws_sum.column_dimensions[get_column_letter(ci)].width = w

    for ri, row_data in enumerate(summary_rows, start=2):
        for ci, col in enumerate(sum_df.columns, start=1):
            val = row_data[col]
            cell = ws_sum.cell(row=ri, column=ci, value=val if val != '' else None)
            if col in int_sum_cols:
                _apply(cell, SUM_INT)
                cell.number_format = INT_FMT
            elif isinstance(val, float):
                _apply(cell, SUM_NUM)
                cell.number_format = NUM_FMT
            else:
                _apply(cell, SUM_TXT)

    # ── Individual run sheets ──────────────────────────────────────────────
    def write_table(ws, start_row, title, df_table, int_col_names=None):
        int_col_names = int_col_names or set()
        n_cols        = len(df_table.columns)
        # Section header (merged)
        ws.merge_cells(
            start_row=start_row, start_column=1,
            end_row=start_row,   end_column=max(n_cols, 1)
        )
        hdr_cell = ws.cell(row=start_row, column=1, value=title)
        _apply(hdr_cell, SEC_HDR)
        start_row += 1
        # Column headers
        for ci, col in enumerate(df_table.columns, start=1):
            cell = ws.cell(row=start_row, column=ci, value=str(col))
            _apply(cell, COL_HDR)
        start_row += 1
        # Data rows
        for ri, row_vals in enumerate(df_table.itertuples(index=False)):
            for ci, val in enumerate(row_vals, start=1):
                col_name = df_table.columns[ci - 1]
                is_nan   = isinstance(val, float) and np.isnan(val)
                write_val = None if (val is None or is_nan) else val
                cell = ws.cell(row=start_row + ri, column=ci, value=write_val)
                if col_name in int_col_names and isinstance(val, (int, float, np.integer, np.floating)) and not is_nan:
                    _apply(cell, RUN_INT)
                    cell.number_format = INT_FMT
                elif isinstance(val, (int, float, np.integer, np.floating)) and not is_nan:
                    _apply(cell, RUN_NUM)
                    cell.number_format = NUM_FMT
                else:
                    _apply(cell, RUN_TXT)
        return start_row + len(df_table) + 2

    for r in run_history:
        is_ridge = r.get("is_ridge_run", False)
        sname    = f"{'R' if is_ridge else 'Run'}{r['run_id']}"
        ws       = wb.create_sheet(sname)
        for ci in range(1, 9):
            ws.column_dimensions[get_column_letter(ci)].width = 22

        rs_df, an_df, cf_df = r['tables']
        row = 1

        if is_ridge:
            # Ridge run: tiêu đề rõ ràng là Ridge
            row = write_table(ws, row,
                f"🔷 RIDGE — Run #{r['run_id']}  ·  {r['dep_var']} ~ {', '.join(r['ind_vars'])}  [{r['model_type']}]",
                rs_df)
            row = write_table(ws, row, "THÔNG TIN RIDGE", an_df)
            row = write_table(ws, row, "HỆ SỐ RIDGE SO VỚI OLS (đã chuẩn hóa X)", cf_df)
        else:
            # OLS run bình thường
            row = write_table(ws, row,
                f"RUN {r['run_id']}  ·  {r['dep_var']} ~ {', '.join(r['ind_vars'])}  [{r['model_type']}]",
                rs_df)
            row = write_table(ws, row, "BẢNG ANOVA", an_df, int_col_names={'df'})
            row = write_table(ws, row, "HỆ SỐ HỒI QUY (COEFFICIENTS)", cf_df)

            # Nếu OLS run này có ridge_result được patch — bỏ qua (đã có sheet Ridge riêng)

        # Interpretation
        hdr_cell = ws.cell(row=row, column=1, value='DIỄN GIẢI KẾT QUẢ')
        _apply(hdr_cell, INTERP_H)
        row += 1
        for i, line in enumerate(r['interpretation'].split('\n')):
            cell = ws.cell(row=row + i, column=1, value=line)
            _apply(cell, INTERP_T)
            ws.row_dimensions[row + i].height = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════════
with st.sidebar:
    # ── Logo ─────────────────────────────────────────────────────────────────
    st.markdown("""
    <div style="display:flex;align-items:center;gap:14px;padding:10px 0 6px 0;">
      <div>
        <svg width="52" height="52" viewBox="0 0 52 52" xmlns="http://www.w3.org/2000/svg">
          <defs>
            <linearGradient id="bg_grad" x1="0%" y1="0%" x2="100%" y2="100%">
              <stop offset="0%" style="stop-color:#1f3a5f;stop-opacity:1"/>
              <stop offset="100%" style="stop-color:#2d1f5f;stop-opacity:1"/>
            </linearGradient>
            <linearGradient id="line_grad" x1="0%" y1="0%" x2="100%" y2="0%">
              <stop offset="0%" style="stop-color:#58a6ff"/>
              <stop offset="50%" style="stop-color:#bc8cff"/>
              <stop offset="100%" style="stop-color:#f778ba"/>
            </linearGradient>
          </defs>
          <!-- Background rounded square -->
          <rect width="52" height="52" rx="13" fill="url(#bg_grad)"/>
          <!-- Grid lines subtle -->
          <line x1="12" y1="40" x2="44" y2="40" stroke="#30363d" stroke-width="0.8"/>
          <line x1="12" y1="30" x2="44" y2="30" stroke="#30363d" stroke-width="0.8"/>
          <line x1="12" y1="20" x2="44" y2="20" stroke="#30363d" stroke-width="0.8"/>
          <line x1="12" y1="40" x2="12" y2="12" stroke="#30363d" stroke-width="0.8"/>
          <!-- Regression line gradient -->
          <line x1="13" y1="38" x2="43" y2="14" stroke="url(#line_grad)" stroke-width="2.5" stroke-linecap="round"/>
          <!-- Scatter dots -->
          <circle cx="16" cy="36" r="2.8" fill="#58a6ff" opacity="0.9"/>
          <circle cx="22" cy="32" r="2.8" fill="#70b8ff" opacity="0.9"/>
          <circle cx="27" cy="27" r="2.8" fill="#9d7aef" opacity="0.9"/>
          <circle cx="33" cy="23" r="2.8" fill="#bc8cff" opacity="0.9"/>
          <circle cx="39" cy="17" r="2.8" fill="#f778ba" opacity="0.9"/>
          <!-- Residual tick lines -->
          <line x1="22" y1="31" x2="22" y2="28.5" stroke="#58a6ff" stroke-width="1.2" opacity="0.6"/>
          <line x1="33" y1="23" x2="33" y2="25.5" stroke="#f778ba" stroke-width="1.2" opacity="0.6"/>
          <!-- Small R² badge -->
          <rect x="34" y="5" width="14" height="9" rx="3" fill="#3fb950" opacity="0.85"/>
          <text x="41" y="12.5" text-anchor="middle" font-family="monospace" font-size="6.5" font-weight="bold" fill="white">R²</text>
        </svg>
      </div>
      <div>
        <div style="font-family:'Space Mono',monospace;font-size:1.15rem;font-weight:700;
             background:linear-gradient(90deg,#58a6ff,#bc8cff,#f778ba);
             -webkit-background-clip:text;-webkit-text-fill-color:transparent;line-height:1.25;">
          Regression<br>Analyst
        </div>
        <div style="color:#8b949e;font-size:0.72rem;margin-top:2px;">OLS · Đa biến · Phi tuyến</div>
      </div>
    </div>
    """, unsafe_allow_html=True)
    st.divider()

    uploaded = st.file_uploader("Upload dữ liệu (.xlsx, .xls, .csv)", type=["xlsx", "xls", "csv"])

    if uploaded:
        try:
            df_loaded, hdr_row = load_data(uploaded)
            if df_loaded is not None:
                st.session_state["df"] = df_loaded
                st.session_state["outlier_checked"] = False
                st.session_state["outlier_rows"] = []
                st.session_state["excluded_rows"] = set()
                st.session_state["filename"] = uploaded.name
                st.session_state["header_row"] = hdr_row
                st.success(f"✅ Đã tải: {uploaded.name}")
                st.caption(f"{df_loaded.shape[0]} hàng × {df_loaded.shape[1]} cột")
                if hdr_row > 0:
                    st.info(f"📌 Tự phát hiện header tại dòng {hdr_row + 1}")
                removed = st.session_state.get("_agg_rows_removed", [])
                if removed:
                    labels = ", ".join(f'"{r}"' for r in removed)
                    st.warning(f"🗑️ Tự động loại {len(removed)} dòng tổng hợp: {labels}")
        except Exception as e:
            st.error(f"Lỗi đọc file: {e}")

    if st.session_state["df"] is not None:
        st.divider()
        if st.button("🗑️ Xóa lịch sử chạy"):
            st.session_state["run_history"] = []
            st.session_state["run_counter"] = 0
            st.rerun()
        if st.session_state["run_history"]:
            st.caption(f"🔢 {len(st.session_state['run_history'])} lần chạy đã lưu")

    # ── Model type guide ──────────────────────────────────────────────────────
    st.divider()
    st.markdown('<div style="font-family:monospace;font-size:0.65rem;text-transform:uppercase;letter-spacing:1.5px;color:#58a6ff;margin-bottom:6px;">📘 Khi nào dùng mô hình nào?</div>', unsafe_allow_html=True)

    with st.expander("📐 Bậc hai Centered — khi nào?", expanded=False):
        st.markdown("""
<div style="font-size:0.82rem;color:#e6edf3;line-height:1.7;">
<b style="color:#58a6ff;">✔ Dùng khi:</b><br>
• Biểu đồ Residual có dạng cong (U-shape hoặc ∩)<br>
• Quan hệ Y~X phi tuyến, cần thêm X²<br>
• Đã thêm X² nhưng hệ số X trở nên không có ý nghĩa (p lớn)<br><br>

<b style="color:#f85149;">⚠ Vấn đề khi dùng Bậc 2 thường:</b><br>
Khi X và X² cùng xuất hiện, chúng thường <b>tương quan rất cao</b> (r ≈ 0.95–0.99) → <b>đa cộng tuyến</b> làm Std Error phồng to, hệ số không ổn định.<br><br>

<b style="color:#3fb950;">✅ Giải pháp — Centering:</b><br>
Xc = X − mean(X)<br>
Xc² = Xc²<br>
→ Loại bỏ tương quan giữa Xc và Xc², hệ số ổn định hơn, <b>R² không đổi</b>.
</div>""", unsafe_allow_html=True)

    with st.expander("📈 Log Y — khi nào?", expanded=False):
        st.markdown("""
<div style="font-size:0.82rem;color:#e6edf3;line-height:1.7;">
<b style="color:#58a6ff;">✔ Dùng khi:</b><br>
• Y tăng theo cấp số nhân (doanh thu, giá, lượt truy cập)<br>
• Phần dư có phương sai tăng dần theo X (heteroscedasticity)<br>
• Biểu đồ Actual vs Predicted bị lệch mạnh với giá trị lớn<br><br>

<b style="color:#d29922;">📌 Lưu ý:</b><br>
• Mô hình: <b>ln(Y) = a + bX</b><br>
• Hệ số b: X tăng 1 đơn vị → Y thay đổi <b>e^b lần</b><br>
• Yêu cầu <b>Y &gt; 0</b> tuyệt đối
</div>""", unsafe_allow_html=True)

    with st.expander("📉 Log X — khi nào?", expanded=False):
        st.markdown("""
<div style="font-size:0.82rem;color:#e6edf3;line-height:1.7;">
<b style="color:#58a6ff;">✔ Dùng khi:</b><br>
• X tăng nhưng ảnh hưởng lên Y <b>giảm dần</b> (diminishing returns)<br>
• Ví dụ: ngân sách quảng cáo tăng → doanh thu tăng, nhưng mỗi đồng thêm vào kém hiệu quả hơn<br>
• Scatter plot Y ~ X có dạng cong lên (logarithm curve), không thẳng<br><br>

<b style="color:#d29922;">📌 Lưu ý:</b><br>
• Mô hình: <b>Y = a + b·ln(X)</b><br>
• Hệ số b: X tăng 1% → Y thay đổi b/100 đơn vị<br>
• Yêu cầu <b>X &gt; 0</b> tuyệt đối<br><br>

<b style="color:#3fb950;">So sánh nhanh:</b><br>
Log Y → dùng khi Y dao động theo cấp số nhân<br>
Log X → dùng khi X tăng nhưng lợi ích giảm dần<br>
Log Y + Log X → dùng khi cả hai (mô hình log-log: hệ số = độ co giãn %)
</div>""", unsafe_allow_html=True)

    with st.expander("🔗 Tương tác (Interaction) — khi nào?", expanded=False):
        st.markdown("""
<div style="font-size:0.82rem;color:#e6edf3;line-height:1.7;">
<b style="color:#58a6ff;">✔ Dùng khi:</b><br>
• Ảnh hưởng của X₁ lên Y <b>phụ thuộc vào mức độ của X₂</b><br>
• Ví dụ: hiệu quả quảng cáo khác nhau theo mùa vụ<br>
• Đã thử mô hình tuyến tính nhưng Q-Q Plot và Residual còn pattern<br><br>

<b style="color:#d29922;">📌 Cách diễn giải:</b><br>
Y = a + b₁X₁ + b₂X₂ + b₃(X₁×X₂)<br>
b₃: khi X₁ tăng 1 đơn vị, hệ số dốc của X₂ thay đổi b₃ đơn vị.<br>
Biểu đồ Trend vẽ <b>2 đường</b> nếu X₂ là biến nhị phân (0/1), hoặc <b>3 đường</b> (−1SD/mean/+1SD) nếu X₂ là biến liên tục — nếu các đường không song song → tương tác có ý nghĩa.<br><br>

<b style="color:#f85149;">⚠ Rủi ro:</b><br>
Dễ gây đa cộng tuyến nếu X₁ và X₂ tương quan cao với nhau.
</div>""", unsafe_allow_html=True)

    with st.expander("🎯 Ridge Regression — khi nào?", expanded=False):
        st.markdown("""
<div style="font-size:0.82rem;color:#e6edf3;line-height:1.7;">
<b style="color:#58a6ff;">✔ Dùng khi:</b><br>
• VIF > 10 — các biến X tương quan cao nhau (đa cộng tuyến)<br>
• OLS cho hệ số không ổn định, p-value lớn dù R² cao<br>
• Dự báo quan trọng hơn diễn giải từng hệ số<br><br>

<b style="color:#d29922;">📐 Ý tưởng cốt lõi:</b><br>
OLS tối thiểu: <b>Σ(y − ŷ)²</b><br>
Ridge tối thiểu: <b>Σ(y − ŷ)² + λΣβ²</b><br>
→ Thêm hình phạt <b>λ</b> (lambda) để <i>co</i> hệ số về 0, tránh phóng đại do đa cộng tuyến.<br><br>

<b style="color:#bc8cff;">🔢 Tham số λ (alpha):</b><br>
• λ = 0 → giống hệt OLS<br>
• λ lớn → hệ số bị co mạnh hơn, bias tăng nhưng variance giảm<br>
• Chọn λ tốt nhất qua <b>Cross-Validation</b><br><br>

<b style="color:#3fb950;">✅ Ưu điểm:</b><br>
• Hệ số ổn định hơn OLS khi có đa cộng tuyến<br>
• Không loại bỏ biến, vẫn giữ tất cả predictors<br>
• Dự báo thường tốt hơn OLS ngoài mẫu<br><br>

<b style="color:#f85149;">⚠ Hạn chế:</b><br>
• Hệ số bị <i>bias</i> — không dùng để diễn giải nhân quả<br>
• Không cung cấp p-value chuẩn như OLS<br>
• Cần chuẩn hóa (standardize) X trước khi dùng
</div>""", unsafe_allow_html=True)

    st.divider()
    st.markdown('<div style="color:#8b949e;font-size:0.72rem;">Powered by statsmodels · OLS · Ridge</div>', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════
# MAIN CONTENT
# ═══════════════════════════════════════════════════════════════════
df = st.session_state["df"]

if df is None:
    st.markdown('<div class="hero-title">📈 Regression Analyst</div>', unsafe_allow_html=True)
    st.markdown('<div class="hero-sub">Tải file dữ liệu (.xlsx, .csv) từ sidebar để bắt đầu phân tích hồi quy</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="card card-accent">
    <b>Tính năng chính:</b><br><br>
    🔍 <b>Kiểm tra chất lượng dữ liệu</b> — phát hiện missing, duplicate, outlier tự động<br>
    📌 <b>Chọn biến linh hoạt</b> — biến phụ thuộc, biến độc lập tùy chọn<br>
    📐 <b>Nhiều loại mô hình</b> — tuyến tính, bậc 2, centered, log, tương tác<br>
    📊 <b>4 biểu đồ mỗi lần chạy</b> — Trend, Residual, Q-Q, Actual vs Predicted<br>
    📋 <b>Bảng so sánh tích lũy</b> — lưu tất cả kết quả để so sánh<br>
    💬 <b>Diễn giải tự động</b> — giải thích hệ số, R², ý nghĩa thống kê<br>
    ⬇️ <b>Xuất Excel chi tiết</b> — đầy đủ bảng ANOVA, hệ số, diễn giải
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ── Section 1: Data Preview & Quality ─────────────────────────────────────
st.markdown('<div class="section-hdr">① DỮ LIỆU & KIỂM TRA CHẤT LƯỢNG</div>', unsafe_allow_html=True)

# Header detection notice
hdr_row = st.session_state.get("header_row", 0)
if hdr_row > 0:
    st.markdown(
        f'<div class="ok-box">✅ <b>Tự động phát hiện tiêu đề:</b> App đã bỏ qua {hdr_row} dòng trên cùng '
        f'và dùng dòng {hdr_row + 1} làm tên cột. Dữ liệu hiển thị bên dưới đã đúng định dạng.</div>',
        unsafe_allow_html=True
    )

col_prev, col_qual = st.columns([1.6, 1])

with col_prev:
    with st.expander("📋 Xem trước dữ liệu (20 dòng đầu)", expanded=True):
        st.dataframe(df.head(20), use_container_width=True, height=270)

with col_qual:
    st.markdown("**Báo cáo chất lượng dữ liệu**")
    issues = data_quality_report(df)
    level_map = {"ok": "ok-box", "warn": "warn-box", "err": "err-box", "info": "info-box"}
    icon_map  = {"ok": "✅", "warn": "⚠️", "err": "🚨", "info": "ℹ️"}
    for iss in issues:
        css  = level_map.get(iss["level"], "info-box")
        icon = icon_map.get(iss["level"], "•")
        st.markdown(f'<div class="{css}">{icon} {iss["msg"]}</div>', unsafe_allow_html=True)

    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    st.caption(f"Kích thước: {df.shape[0]} hàng × {df.shape[1]} cột")
    st.caption(f"Cột số: {len(numeric_cols)} | Cột phi số: {df.shape[1] - len(numeric_cols)}")

st.divider()

# ── Section 2: Model Configuration ────────────────────────────────────────
st.markdown('<div class="section-hdr">② CẤU HÌNH MÔ HÌNH HỒI QUY</div>', unsafe_allow_html=True)

if len(numeric_cols) < 2:
    st.error("Cần ít nhất 2 cột số để thực hiện hồi quy.")
    st.stop()

# ── Smart variable suggestion ─────────────────────────────────────────────
sug_y, sug_x, _, sug_text = suggest_variables(df)

with st.expander("🤖 Gợi ý biến tự động (click để xem / ẩn)", expanded=True):
    st.markdown(sug_text)
    # Correlation heatmap mini-table: show top correlations with suggested Y
    if sug_y and len(numeric_cols) >= 2:
        corr_series = df[numeric_cols].corr()[sug_y].drop(sug_y).sort_values(key=abs, ascending=False)
        corr_df = corr_series.reset_index()
        corr_df.columns = ["Biến X", f"Tương quan với {sug_y}"]
        corr_df[f"Tương quan với {sug_y}"] = corr_df[f"Tương quan với {sug_y}"].round(4)

        def color_corr(val):
            try:
                v = float(val)
                if abs(v) >= 0.7:
                    return "background-color:#1a3a2a; color:#3fb950"
                elif abs(v) >= 0.4:
                    return "background-color:#3a2d10; color:#d29922"
                else:
                    return "background-color:#2a1f1f; color:#8b949e"
            except Exception:
                return ""

        try:
            styled = corr_df.style.map(color_corr, subset=[f"Tương quan với {sug_y}"])
        except AttributeError:
            styled = corr_df.style.applymap(color_corr, subset=[f"Tương quan với {sug_y}"])
        st.dataframe(styled, use_container_width=True, hide_index=True, height=min(200, 35 * len(corr_df) + 38))
        st.caption("🟢 |r| ≥ 0.7 mạnh · 🟡 0.4–0.7 trung bình · ⬜ < 0.4 yếu")

st.markdown("")  # spacing

# ── Variable selectors (pre-filled with suggestion) ───────────────────────
col_cfg1, col_cfg2, col_cfg3 = st.columns([1, 1.2, 1])

# Find suggested Y index for selectbox default
sug_y_idx = numeric_cols.index(sug_y) if sug_y in numeric_cols else 0

with col_cfg1:
    dep_var = st.selectbox(
        "🎯 Biến phụ thuộc (Y)",
        options=numeric_cols,
        index=sug_y_idx,
        help="Biến bạn muốn dự báo / giải thích. Đã được gợi ý tự động — có thể thay đổi."
    )

with col_cfg2:
    ind_options = [c for c in numeric_cols if c != dep_var]
    # Use suggested X, filtered to only valid options
    default_x = [c for c in sug_x if c in ind_options] or (ind_options[:1] if ind_options else [])
    ind_vars = st.multiselect(
        "📌 Biến độc lập (X)",
        options=ind_options,
        default=default_x,
        help="Chọn một hoặc nhiều biến giải thích. Đã gợi ý tự động — có thể thêm/bỏ."
    )

with col_cfg3:
    model_type = st.selectbox(
        "📐 Loại mô hình",
        options=[
            "Tuyến tính (Linear)",
            "Bậc hai (Quadratic)",
            "Bậc hai Centered",
            "Logarithmic (Log Y)",
            "Logarithmic (Log X)",
            "Tương tác (Interaction)"
        ],
        help=(
            "Linear: Y = a + b₁X₁ + … | "
            "Quadratic: + bX² | "
            "Centered: giảm đa cộng tuyến | "
            "Log Y: ln(Y) = … | "
            "Log X: Y = a + b·ln(X) | "
            "Interaction: thêm tích X₁×X₂"
        )
    )

# Model type info box
MODEL_INFO = {
    "Tuyến tính (Linear)":     "Hồi quy tuyến tính chuẩn. Phù hợp khi quan hệ Y~X là đường thẳng.",
    "Bậc hai (Quadratic)":     "Thêm X² vào mô hình — phù hợp khi đồ thị phần dư có dạng cong (U-shape).",
    "Bậc hai Centered":        "Bậc 2 với X được trừ giá trị trung bình — giảm đa cộng tuyến giữa X và X².",
    "Logarithmic (Log Y)":     "Biến đổi ln(Y) trước khi hồi quy — phù hợp khi Y tăng theo cấp số nhân (Y > 0 bắt buộc).",
    "Logarithmic (Log X)":     "Dùng ln(X) làm biến giải thích — phù hợp khi X tăng nhưng ảnh hưởng đến Y giảm dần (diminishing returns). X > 0 bắt buộc.",
    "Tương tác (Interaction)": "Thêm tích X₁×X₂ — phù hợp khi ảnh hưởng của X₁ phụ thuộc vào X₂. Nếu X₂ là nhị phân (0/1): vẽ 2 đường. Nếu X₂ liên tục: vẽ 3 đường (−1SD/mean/+1SD).",
}
st.markdown(f'<div class="info-box">ℹ️ <b>{model_type}</b>: {MODEL_INFO[model_type]}</div>', unsafe_allow_html=True)

st.divider()

# ── Section 3: Outlier Warning & Run Model ──────────────────────────────────
st.markdown('<div class="section-hdr">③ CẢNH BÁO OUTLIER & CHẠY MÔ HÌNH</div>', unsafe_allow_html=True)

# ── Outlier detection (runs once per loaded dataframe) ─────────────────────
df_id = id(df)
if not st.session_state.get("outlier_checked") or st.session_state.get("_df_id") != df_id:
    st.session_state["outlier_rows"] = detect_outlier_rows(df)
    st.session_state["excluded_rows"] = set()
    st.session_state["outlier_checked"] = True
    st.session_state["_df_id"] = df_id

outlier_rows = st.session_state["outlier_rows"]
excluded_rows = st.session_state["excluded_rows"]

if outlier_rows:
    st.markdown('<div class="warn-box">⚠️ <b>Phát hiện dữ liệu bất thường!</b> App đã tìm thấy các dòng dưới đây có thể ảnh hưởng đến kết quả hồi quy. Vui lòng xem xét và chọn có loại bỏ hay không trước khi chạy.</div>', unsafe_allow_html=True)

    with st.expander("🔍 Chi tiết dòng bất thường — click để xem và chọn", expanded=True):
        st.markdown("**Đánh dấu ✗ để loại dòng, để trống để giữ lại:**")
        new_excluded = set()
        for item in outlier_rows:
            idx = item["index"]
            severity_icon = "🚫" if item["severity"] == "aggregate" else "⚠️"
            label_text = f"{severity_icon} **Dòng {idx}** — {item['label']}  |  {item['reason']}"
            checked = st.checkbox(label_text, value=(idx in excluded_rows), key=f"excl_{idx}")
            if checked:
                new_excluded.add(idx)
        st.session_state["excluded_rows"] = new_excluded
        excluded_rows = new_excluded

    if excluded_rows:
        st.markdown(
            f'<div class="info-box">ℹ️ Đã chọn loại <b>{len(excluded_rows)}</b> dòng: index {sorted(excluded_rows)}. ' +
            'App sẽ chạy <b>2 lần</b>: một lần có đầy đủ dữ liệu và một lần đã loại, để so sánh.</div>',
            unsafe_allow_html=True
        )
    else:
        st.markdown('<div class="info-box">ℹ️ Chưa chọn loại dòng nào — sẽ chạy với toàn bộ dữ liệu.</div>', unsafe_allow_html=True)
else:
    st.markdown('<div class="ok-box">✅ Không phát hiện dòng tổng hợp hay outlier cực đoan trong dữ liệu.</div>', unsafe_allow_html=True)

run_col, _ = st.columns([1, 3])
with run_col:
    run_clicked = st.button("▶ Chạy Hồi Quy", type="primary", use_container_width=True)

def _do_one_run(df_run, dep_var, ind_vars, model_type, label_suffix=""):
    """Run one regression and return a result dict, or raise."""
    model, df_clean, y = run_regression(df_run, dep_var, ind_vars, model_type)
    rs_df, an_df, cf_df = get_ols_stats(model)
    interp = interpret_model(model, model_type, dep_var, ind_vars)
    plot_buf = make_plots(model, df_clean, dep_var, ind_vars, model_type)
    st.session_state["run_counter"] += 1
    run_id = st.session_state["run_counter"]
    return {
        "run_id": run_id,
        "dep_var": dep_var,
        "ind_vars": ind_vars.copy(),
        "model_type": model_type,
        "n": int(model.nobs),
        "r2": round(model.rsquared, 4),
        "adj_r2": round(model.rsquared_adj, 4),
        "rmse": round(float(np.sqrt(model.mse_resid)), 4),
        "fstat": round(model.fvalue, 4),
        "f_pvalue": round(model.f_pvalue, 6),
        "tables": (rs_df, an_df, cf_df),
        "interpretation": interp,
        "plot_buf": plot_buf,
        "label_suffix": label_suffix,
        "params": dict(model.params),         # for Ridge coef comparison
        "ridge_result": None,                  # filled when Ridge is run
    }

if run_clicked:
    if not ind_vars:
        st.error("Vui lòng chọn ít nhất 1 biến độc lập.")
    else:
        with st.spinner("Đang chạy mô hình..."):
            try:
                # Always run with full data first
                result_full = _do_one_run(df, dep_var, ind_vars, model_type, " [Đầy đủ]")
                st.session_state["run_history"].append(result_full)
                st.success(f"✅ Run #{result_full['run_id']} (đầy đủ dữ liệu) — R² = {result_full['r2']:.4f}")

                # If user excluded rows, also run without them
                if excluded_rows:
                    df_filtered = df.drop(index=list(excluded_rows)).reset_index(drop=True)
                    result_excl = _do_one_run(df_filtered, dep_var, ind_vars, model_type, " [Đã loại outlier]")
                    st.session_state["run_history"].append(result_excl)
                    st.success(f"✅ Run #{result_excl['run_id']} (đã loại {len(excluded_rows)} dòng) — R² = {result_excl['r2']:.4f}")

                    # Quick comparison callout
                    delta_r2 = result_excl["r2"] - result_full["r2"]
                    delta_rmse = result_excl["rmse"] - result_full["rmse"]
                    arrow_r2   = "↑" if delta_r2 > 0 else "↓"
                    arrow_rmse = "↓" if delta_rmse < 0 else "↑"
                    box_cls = "ok-box" if delta_r2 > 0 else "warn-box"
                    st.markdown(
                        f'<div class="{box_cls}">📊 <b>So sánh nhanh sau khi loại outlier:</b> ' +
                        f'R² {arrow_r2} {abs(delta_r2):.4f} (đầy đủ: {result_full["r2"]:.4f} → loại: {result_excl["r2"]:.4f}) | ' +
                        f'RMSE {arrow_rmse} {abs(delta_rmse):.4f} (đầy đủ: {result_full["rmse"]:.4f} → loại: {result_excl["rmse"]:.4f})</div>',
                        unsafe_allow_html=True
                    )
            except Exception as e:
                st.error(f"Lỗi khi chạy mô hình: {e}")


# ── Section 4: Latest Run Results ──────────────────────────────────────────
if st.session_state["run_history"]:
    latest = st.session_state["run_history"][-1]

    suffix = latest.get("label_suffix", "")
    st.markdown(f'<div class="section-hdr">④ KẾT QUẢ MÔ HÌNH — Run #{latest["run_id"]}{suffix}: {latest["dep_var"]} ~ {" + ".join(latest["ind_vars"])} [{latest["model_type"]}]</div>', unsafe_allow_html=True)

    # Metrics row
    mc1, mc2, mc3, mc4, mc5 = st.columns(5)
    mc1.metric("R²", f"{latest['r2']:.4f}")
    mc2.metric("Adj R²", f"{latest['adj_r2']:.4f}")
    mc3.metric("RMSE", f"{latest['rmse']:.4f}")
    mc4.metric("F-stat", f"{latest['fstat']:.2f}")
    mc5.metric("F p-value", f"{latest['f_pvalue']:.4f}")

    # Tables
    tab1, tab2, tab3 = st.tabs(["📊 Regression Statistics", "📋 ANOVA", "🔢 Coefficients"])
    rs_df, an_df, cf_df = latest["tables"]
    with tab1:
        st.dataframe(rs_df, use_container_width=True, hide_index=True)
    with tab2:
        st.dataframe(an_df, use_container_width=True, hide_index=True)
    with tab3:
        st.dataframe(cf_df, use_container_width=True, hide_index=True)

    # ── VIF / Multicollinearity check ────────────────────────────────────────
    if len(latest["ind_vars"]) >= 2:
        try:
            df_for_vif = df[[latest["dep_var"]] + latest["ind_vars"]].dropna().copy()
            # For Quadratic / Centered models, rebuild original numeric vars
            if latest["model_type"] in ["Bậc hai (Quadratic)", "Bậc hai Centered", "Tương tác (Interaction)"]:
                # Use the raw ind_vars that were passed in (before squaring)
                vif_df, has_high_vif, poly_col = compute_vif(df, latest["ind_vars"])
            else:
                vif_df, has_high_vif, poly_col = compute_vif(df, latest["ind_vars"])

            st.markdown("**🔬 Kiểm tra đa cộng tuyến (VIF)**")
            vif_cols = st.columns(len(vif_df))
            for i, row_v in vif_df.iterrows():
                color = row_v["_color"]
                vif_cols[i % len(vif_cols)].markdown(
                    f'''<div style="background:#1c2333;border:1px solid {color};border-radius:8px;
                    padding:8px 12px;text-align:center;margin-bottom:4px;">
                    <div style="color:#8b949e;font-size:0.75rem;">{row_v["Biến"]}</div>
                    <div style="color:{color};font-size:1.3rem;font-weight:700;">{row_v["VIF"]}</div>
                    <div style="color:{color};font-size:0.7rem;">{row_v["Mức độ"]}</div>
                    </div>''', unsafe_allow_html=True
                )

            if has_high_vif:
                is_quad_model = latest["model_type"] in ["Bậc hai (Quadratic)"]
                if poly_col or is_quad_model:
                    st.markdown('''<div class="err-box">
🚨 <b>Đa cộng tuyến cao — nguyên nhân có thể do mô hình bậc 2 (X và X²)!</b><br><br>
<b>Giải thích:</b> Khi thêm biến X² vào mô hình, X và X² thường tương quan rất cao (r ≈ 0.95–0.99) 
làm <b>VIF tăng vọt</b>, Std Error phồng to, và hệ số X mất ý nghĩa thống kê — dù R² vẫn cao.<br><br>
<b>✅ Giải pháp được khuyến nghị — Centering:</b><br>
① Tính biến mới: <code>Xc = X − mean(X)</code> và <code>Xc² = Xc²</code><br>
② Dùng <b>Xc</b> và <b>Xc²</b> thay cho X và X² trong mô hình<br>
③ Chọn mô hình <b>"Bậc hai Centered"</b> trong phần Cấu hình → R² giống hệt nhưng VIF giảm mạnh<br><br>
<b>Lưu ý:</b> Centering không thay đổi khả năng dự báo, chỉ làm hệ số ổn định và dễ diễn giải hơn.
</div>''', unsafe_allow_html=True)
                else:
                    st.markdown('''<div class="warn-box">
⚠️ <b>Đa cộng tuyến cao giữa các biến độc lập!</b><br>
VIF > 10 — các biến X tương quan cao với nhau làm hệ số hồi quy không ổn định, Std Error phồng to.<br>
Chọn một trong 3 giải pháp bên dưới:
</div>''', unsafe_allow_html=True)

                    sol1, sol2, sol3 = st.tabs(["① Loại biến thừa", "② Ridge Regression", "③ Xem lại đặc tả"])

                    with sol1:
                        st.markdown('''<div class="info-box" style="font-size:0.85rem;line-height:1.8;">
<b>Loại biến có ý nghĩa thống kê thấp nhất:</b><br>
① Xem bảng <b>Coefficients</b> → tìm biến có <b>p-value lớn nhất</b> (> 0.05)<br>
② Loại biến đó khỏi danh sách X trong phần Cấu hình<br>
③ Chạy lại mô hình và kiểm tra VIF — lặp cho đến khi VIF < 10<br><br>
<b>Lưu ý:</b> Chỉ loại nếu biến đó thực sự không có lý thuyết hỗ trợ. Đừng loại thuần túy vì p-value.
</div>''', unsafe_allow_html=True)

                    with sol2:
                        if not _SKLEARN_OK:
                            st.error("⚠️ scikit-learn chưa được cài. Thêm `scikit-learn` vào requirements.txt rồi restart app.")
                        else:
                            st.markdown('''<div class="info-box" style="font-size:0.85rem;line-height:1.8;">
<b>Ridge Regression</b> thêm hình phạt λΣβ² vào OLS — co hệ số về 0 để ổn định hóa ước lượng khi có đa cộng tuyến.<br>
Dữ liệu X được <b>chuẩn hóa (StandardScaler)</b> trước khi fit — hệ số Ridge cho phép so sánh tầm quan trọng tương đối giữa các biến.<br>
App sẽ tự động tìm <b>λ tối ưu</b> bằng 5-fold Cross-Validation, sau đó chạy Ridge với λ đó.
</div>''', unsafe_allow_html=True)

                            if st.button("🔍 Tìm λ tối ưu & Chạy Ridge", key=f"ridge_run_{latest['run_id']}"):
                                try:
                                    df_r = df[latest["ind_vars"] + [latest["dep_var"]]].dropna().copy()
                                    X_r  = df_r[latest["ind_vars"]].values.astype(float)
                                    y_r  = df_r[latest["dep_var"]].values.astype(float)

                                    scaler     = StandardScaler()
                                    X_scaled   = scaler.fit_transform(X_r)

                                    # ── Step 1: find optimal λ ──────────────────────────────
                                    alphas_cv  = np.logspace(-3, 2, 100)   # 0.001 → 100
                                    ridge_cv   = RidgeCV(alphas=alphas_cv, cv=5)
                                    ridge_cv.fit(X_scaled, y_r)
                                    best_lam   = float(ridge_cv.alpha_)

                                    # ── Step 2: fit Ridge with best λ ───────────────────────
                                    ridge      = Ridge(alpha=best_lam)
                                    ridge.fit(X_scaled, y_r)
                                    y_pred_r   = ridge.predict(X_scaled)

                                    r2_ridge   = float(_sk_r2(y_r, y_pred_r))
                                    rmse_ridge = float(np.sqrt(np.mean((y_r - y_pred_r) ** 2)))
                                    ols_r2     = latest["r2"]
                                    delta      = r2_ridge - ols_r2
                                    dir_str    = f"giảm {abs(delta):.4f}" if delta < 0 else f"tăng {abs(delta):.4f}"

                                    st.markdown(f'''<div class="ok-box">
✅ <b>Ridge hoàn thành</b> — λ tối ưu (5-fold CV) = <b>{best_lam:.4f}</b><br>
R² Ridge = <b>{r2_ridge:.4f}</b> (so với OLS: {dir_str}) &nbsp;|&nbsp; RMSE = <b>{rmse_ridge:.4f}</b>
</div>''', unsafe_allow_html=True)

                                    # ── Step 3: OLS on scaled data for fair comparison ───────
                                    _X_s        = sm.add_constant(X_scaled)
                                    _ols_s      = sm.OLS(y_r, _X_s).fit()
                                    ols_std     = list(_ols_s.params[1:])   # skip const

                                    ridge_coefs = list(ridge.coef_)
                                    shrinkage   = []
                                    for oc, rc in zip(ols_std, ridge_coefs):
                                        shrinkage.append(
                                            f"{abs(rc)/abs(oc)*100:.1f}%" if abs(oc) > 1e-9 else "—"
                                        )

                                    coef_df = pd.DataFrame({
                                        "Biến":               latest["ind_vars"],
                                        "Hệ số OLS (std)":   [round(c, 4) for c in ols_std],
                                        "Hệ số Ridge (std)":  [round(c, 4) for c in ridge_coefs],
                                        "Co lại còn (%)":     shrinkage,
                                    })
                                    st.dataframe(coef_df, use_container_width=True, hide_index=True)

                                    # ── Step 4: Interpretation guide ────────────────────────
                                    # Find most-shrunk variable
                                    shrink_pct = []
                                    for oc, rc in zip(ols_std, ridge_coefs):
                                        try:
                                            shrink_pct.append(abs(rc)/abs(oc)*100 if abs(oc) > 1e-9 else 100.0)
                                        except Exception:
                                            shrink_pct.append(100.0)
                                    most_shrunk_idx = int(np.argmin(shrink_pct))
                                    most_shrunk_var = latest["ind_vars"][most_shrunk_idx]
                                    most_shrunk_pct = shrink_pct[most_shrunk_idx]

                                    # Detect sign flip (strong multicollinearity signal)
                                    sign_flips = [
                                        v for v, oc, rc in zip(latest["ind_vars"], ols_std, ridge_coefs)
                                        if np.sign(oc) != np.sign(rc) and abs(oc) > 0.01
                                    ]
                                    sign_flip_html = ""
                                    if sign_flips:
                                        sign_flip_html = f"<br><b>⚠ Biến đổi dấu:</b> {', '.join(sign_flips)} — Trong OLS, đa cộng tuyến làm hệ số bị đảo dấu. Ridge đã sửa lại chiều tác động đúng hơn."

                                    st.markdown(f'''<div class="info-box" style="font-size:0.82rem;line-height:1.8;">
<b>🔍 Đa cộng tuyến được khắc phục như thế nào?</b><br><br>

<b>① Cột "Co lại còn (%)":</b> Biến nào có % thấp = bị ảnh hưởng đa cộng tuyến nhiều nhất trong OLS.<br>
Biến bị co mạnh nhất: <b>{most_shrunk_var}</b> (còn {most_shrunk_pct:.1f}% so với OLS). Ridge làm ổn định hệ số này.{sign_flip_html}<br><br>

<b>② Ridge không làm giảm VIF</b> — VIF đo tương quan giữa các X, Ridge không thay đổi dữ liệu X. Ridge kiểm soát <i>hậu quả</i>: hệ số không còn dao động mạnh giữa các mẫu.<br><br>

<b>③ R² Ridge {dir_str} so với OLS</b> — đây là bình thường. Ridge hi sinh một chút độ khớp trong mẫu để đổi lấy hệ số ổn định hơn khi dự báo dữ liệu mới.<br><br>

<b>④ Khi nào Ridge thực sự giúp ích?</b> Khi VIF cao (>10) và p-value của các biến quan trọng bị thổi phồng — Ridge ổn định hệ số để diễn giải và dự báo đáng tin hơn.
</div>''', unsafe_allow_html=True)

                                    # ── Step 5: Lưu ridge_result vào run gốc (để Excel có đủ)
                                    ridge_summary = {
                                        "best_lambda":   round(best_lam, 4),
                                        "r2":            round(r2_ridge, 4),
                                        "rmse":          round(rmse_ridge, 4),
                                        "coef_df":       coef_df,
                                        "most_shrunk":   most_shrunk_var,
                                    }
                                    st.session_state["run_history"][-1]["ridge_result"] = ridge_summary

                                    # ── Step 6: Thêm run MỚI cho Ridge vào run_history ──────
                                    st.session_state["run_counter"] += 1
                                    ridge_run_id = st.session_state["run_counter"]

                                    # Tạo interpretation riêng cho Ridge
                                    ridge_interp_lines = [
                                        f"📊 RIDGE REGRESSION — Từ Run #{latest['run_id']} (OLS)",
                                        f"• λ tối ưu (5-fold CV) = {round(best_lam, 4)}",
                                        f"• R² = {round(r2_ridge, 4)} (OLS: {latest['r2']}, delta: {round(r2_ridge - latest['r2'], 4):+.4f})",
                                        f"• RMSE = {round(rmse_ridge, 4)} (OLS: {latest['rmse']})",
                                        "",
                                        "🔢 HỆ SỐ RIDGE (đã chuẩn hóa X)",
                                    ]
                                    for _, row_r in coef_df.iterrows():
                                        ridge_interp_lines.append(
                                            f"• {row_r['Biến']}: OLS={row_r['Hệ số OLS (std)']}, Ridge={row_r['Hệ số Ridge (std)']}, Co lại={row_r['Co lại còn (%)']}"
                                        )
                                    ridge_interp_lines += [
                                        "",
                                        f"🔍 Biến bị co mạnh nhất: {most_shrunk_var} ({most_shrunk_pct:.1f}% còn lại so với OLS)",
                                        "• Ridge không giảm VIF nhưng ổn định hệ số — phù hợp khi mục tiêu là dự báo.",
                                    ]
                                    if sign_flips:
                                        ridge_interp_lines.append(f"• Biến đổi dấu trong OLS đã được Ridge sửa: {', '.join(sign_flips)}")

                                    # Tạo bảng giả để hiển thị trong Excel sheet Ridge
                                    ridge_rs_df = pd.DataFrame({
                                        "Chỉ số": ["λ (alpha)", "R² Ridge", "RMSE Ridge", "R² OLS gốc", "RMSE OLS gốc", "ΔR²"],
                                        "Giá trị": [
                                            round(best_lam, 4), round(r2_ridge, 4), round(rmse_ridge, 4),
                                            latest["r2"], latest["rmse"], round(r2_ridge - latest["r2"], 4)
                                        ]
                                    })
                                    ridge_an_df = pd.DataFrame({
                                        "Thông tin": ["Phương pháp", "Cross-Validation", "Chuẩn hóa X", "Run OLS tham chiếu"],
                                        "Giá trị":   ["Ridge Regression (L2)", "5-fold CV", "StandardScaler", f"Run #{latest['run_id']}"]
                                    })

                                    ridge_run_entry = {
                                        "run_id":        ridge_run_id,
                                        "dep_var":       latest["dep_var"],
                                        "ind_vars":      latest["ind_vars"].copy(),
                                        "model_type":    f"Ridge (λ={round(best_lam,4)}) ← Run#{latest['run_id']}",
                                        "n":             int(len(y_r)),
                                        "r2":            round(r2_ridge, 4),
                                        "adj_r2":        round(r2_ridge, 4),   # Ridge không có Adj R² chuẩn
                                        "rmse":          round(rmse_ridge, 4),
                                        "fstat":         float("nan"),
                                        "f_pvalue":      float("nan"),
                                        "tables":        (ridge_rs_df, ridge_an_df, coef_df),
                                        "interpretation": "\n".join(ridge_interp_lines),
                                        "plot_buf":      latest["plot_buf"],   # tái dùng biểu đồ OLS gốc
                                        "label_suffix":  f" [Ridge từ Run#{latest['run_id']}]",
                                        "params":        {},
                                        "ridge_result":  ridge_summary,        # giữ để bảng so sánh dùng
                                        "is_ridge_run":  True,                 # flag phân biệt
                                    }
                                    st.session_state["run_history"].append(ridge_run_entry)
                                    st.success(f"✅ Đã lưu Run #{ridge_run_id} (Ridge) vào bảng So Sánh bên dưới.")

                                except Exception as ridge_err:
                                    st.error(f"Lỗi Ridge: {ridge_err}")

                    with sol3:
                        st.markdown('''<div class="info-box" style="font-size:0.85rem;line-height:1.8;">
<b>Kiểm tra lại cách đặc tả mô hình:</b><br>
• Các biến X có <b>đo lường cùng một khái niệm</b> không? (ví dụ: diện tích m² và diện tích ft²)<br>
• Có biến nào là <b>tổ hợp tuyến tính</b> của biến khác không?<br>
• Nếu có biến phân loại dummy, đã <b>loại bỏ 1 category gốc</b> chưa? (dummy trap)<br>
• Thử dùng <b>PCA</b> để giảm chiều trước khi hồi quy (Principal Component Regression)<br><br>
<b>Gợi ý:</b> Kiểm tra ma trận tương quan giữa các biến X bên dưới:
</div>''', unsafe_allow_html=True)
                        try:
                            corr_x = df[latest["ind_vars"]].corr().round(3)
                            st.dataframe(corr_x.style.background_gradient(cmap="RdYlGn_r", vmin=-1, vmax=1),
                                         use_container_width=True)
                        except Exception:
                            st.caption("Không tính được ma trận tương quan.")
            elif len(vif_df) > 0:
                st.markdown('<div class="ok-box">✅ VIF của tất cả biến đều trong ngưỡng an toàn (< 5) — không có đa cộng tuyến đáng lo.</div>', unsafe_allow_html=True)
        except Exception as vif_err:
            st.caption(f"(Không tính được VIF: {vif_err})")

    # Charts
    st.markdown("**📈 Biểu đồ phân tích**")
    latest["plot_buf"].seek(0)
    st.image(latest["plot_buf"], use_container_width=True)

    # Interpretation
    st.markdown("**💬 Diễn giải kết quả**")
    st.markdown(f'<div class="interpret-box">{latest["interpretation"].replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)

    st.divider()


# ── Section 5: Run History Comparison ──────────────────────────────────────
if len(st.session_state["run_history"]) > 0:
    st.markdown('<div class="section-hdr">⑤ SO SÁNH CÁC MÔ HÌNH ĐÃ CHẠY</div>', unsafe_allow_html=True)

    history = st.session_state["run_history"]
    compare_rows = []
    for r in history:
        rr        = r.get("ridge_result")
        is_ridge  = r.get("is_ridge_run", False)
        fstat_val = "—" if is_ridge or (isinstance(r["fstat"], float) and np.isnan(r["fstat"])) else r["fstat"]
        fp_val    = "—" if is_ridge or (isinstance(r["f_pvalue"], float) and np.isnan(r["f_pvalue"])) else r["f_pvalue"]
        compare_rows.append({
            "Run #":            r["run_id"],
            "Loại":             "🔷 Ridge" if is_ridge else "📊 OLS",
            "Ghi chú":          r.get("label_suffix", "").strip(" []"),
            "Y":                r["dep_var"],
            "X (biến độc lập)": ", ".join(r["ind_vars"]),
            "Mô hình":          r["model_type"],
            "N":                r["n"],
            "R²":               r["r2"],
            "Adj R²":           r["adj_r2"] if not is_ridge else "—",
            "RMSE":             r["rmse"],
            "F-stat":           fstat_val,
            "F p-value":        fp_val,
            "Ridge λ":          rr["best_lambda"] if rr else "—",
        })

    cmp_df = pd.DataFrame(compare_rows)
    st.dataframe(cmp_df, use_container_width=True, hide_index=True)

    # Best OLS model recommendation (exclude Ridge runs from ranking)
    ols_history = [r for r in history if not r.get("is_ridge_run", False)]
    if len(ols_history) > 1:
        ols_rows = [row for row in compare_rows if row["Loại"] == "📊 OLS"]
        ols_cmp  = pd.DataFrame(ols_rows)
        best_idx  = ols_cmp["Adj R²"].astype(float).idxmax()
        best      = ols_cmp.iloc[best_idx]
        st.markdown(f"""
        <div class="ok-box">
        🏆 <b>Mô hình OLS tốt nhất theo Adjusted R²:</b> Run #{int(best['Run #'])} — {best['Mô hình']}
        với Y = {best['Y']}, X = {best['X (biến độc lập)']}<br>
        Adj R² = <b>{best['Adj R²']:.4f}</b>, RMSE = <b>{best['RMSE']:.4f}</b>
        </div>
        """, unsafe_allow_html=True)

        min_rmse_idx = ols_cmp["RMSE"].astype(float).idxmin()
        if min_rmse_idx != best_idx:
            best_rmse = ols_cmp.iloc[min_rmse_idx]
            st.markdown(f"""
            <div class="info-box">
            ℹ️ <b>RMSE thấp nhất (OLS):</b> Run #{int(best_rmse['Run #'])} — {best_rmse['Mô hình']}
            (RMSE = {best_rmse['RMSE']:.4f}).
            Nếu mục tiêu là tối thiểu sai số dự báo, hãy xem xét mô hình này.
            </div>
            """, unsafe_allow_html=True)

    # Detailed history expander
    if len(history) > 1:
        with st.expander("🔍 Xem chi tiết từng lần chạy"):
            for r in history:
                is_ridge = r.get("is_ridge_run", False)
                icon     = "🔷" if is_ridge else "📊"
                st.markdown(f"**{icon} Run #{r['run_id']}: {r['dep_var']} ~ {', '.join(r['ind_vars'])} [{r['model_type']}]**")
                _, _, cf = r["tables"]
                st.dataframe(cf, use_container_width=True, hide_index=True)
                if is_ridge:
                    rr = r.get("ridge_result")
                    st.caption(f"R² = {r['r2']} | RMSE = {r['rmse']} | λ = {rr['best_lambda'] if rr else '—'}")
                else:
                    st.caption(f"R² = {r['r2']} | Adj R² = {r['adj_r2']} | RMSE = {r['rmse']}")
                st.divider()

    st.divider()

    # ── Export ─────────────────────────────────────────────────────────────
    st.markdown('<div class="section-hdr">⑥ XUẤT KẾT QUẢ EXCEL</div>', unsafe_allow_html=True)

    excel_col, info_col = st.columns([1, 2])
    with excel_col:
        excel_bytes = export_to_excel(st.session_state["run_history"])
        fname = f"RegressionResults_{st.session_state['filename'].split('.')[0]}.xlsx"
        st.download_button(
            label="⬇️ Tải xuống Excel (tất cả lần chạy)",
            data=excel_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with info_col:
        st.markdown("""
        <div class="info-box">
        File Excel bao gồm:<br>
        • <b>Sheet Summary</b>: bảng so sánh tất cả lần chạy<br>
        • <b>Sheet Run#N</b>: Regression Statistics + ANOVA + Coefficients + Diễn giải cho từng lần chạy
        </div>
        """, unsafe_allow_html=True)
